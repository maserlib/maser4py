#!/usr/bin/env python3
# -*- coding: utf-8 -*-

# ________________ HEADER _________________________


"""Skt2xlsx module

Program to convert an input CDF skeleton table file
into an Excel format file (skt_editor).

"""


# ________________ IMPORT _________________________
# (Include here the modules to import, e.g. import sys)
import os
import logging
from collections import OrderedDict
from copy import deepcopy

from openpyxl import Workbook, load_workbook

from maser.utils.toolbox import uniq
from maser.utils.cdf.serializer.exceptions import InvalidFile, InvalidEntry
from maser.utils.cdf.serializer.globals import SHEETS, HEADER, GATTRS, ZVARS, VATTRS, NRV

__all__ = ['Skt2xlsx', 'Xlsx2skt']

# ________________ Global Variables _____________
# (define here the global variables)
logger = logging.getLogger(__name__)

# ________________ Class Definition __________
# (If required, define here classes)


class Skt2xlsx:
    """Class to convert a Skeleton object into a Excel 2007 format file."""

    def __init__(self, skeleton,
                 auto_pad=True):
        """__init__ method."""
        self.skeleton = deepcopy(skeleton)
        self.wb = None
        self.auto_pad = auto_pad

    def write_xlsx(self, output_path=None,
                   overwrite=False):
        """
        Write the Excel 2007 format file.

        :param output_path:
        :param overwrite: If true, overwrite existing output file
        :return:
        """

        if not output_path:
            xlsx = os.path.splitext(self.skeleton.file)[0] + '.xlsx'
        elif os.path.isdir(output_path):
            xlsx = os.path.join(output_path,
                                os.path.basename(
                                    os.path.splitext(self.skeleton.file)[
                                        0] + '.xlsx'
                                ))
        else:
            xlsx = output_path

        if not overwrite and os.path.isfile(xlsx):
            logger.warning('%s already exits!', xlsx)
            return xlsx

        logger.info('Writing %s...', xlsx)

        if not self.wb:
            logger.debug('Attempting to build Excel file body...')
            self._build_wb()

        self.wb.save(filename=xlsx)

        if os.path.isfile(xlsx):
            logger.info('{0} has been saved correctly'.format(xlsx))
            return xlsx
        else:
            logger.error('{0} has not been saved correctly!'.format(xlsx))
            return None

    def _build_wb(self):
        """
        Create Workbook from Skeleton object.

        :return:
        """
        wb = Workbook()

        # Create header sheet
        header_sheet = wb.create_sheet(title=HEADER)

        col_idx = 1
        for key, val in self.skeleton.header.items():
            if key not in SHEETS[HEADER]:
                continue
            # Make header of header in first row
            _ = header_sheet.cell(column=col_idx, row=1,
                                  value='{0}'.format(key))
            # Add values in the second row
            _ = header_sheet.cell(column=col_idx, row=2,
                                  value='{0}'.format(val))
            col_idx += 1

        # Create GLOBALattributes sheet
        gattrs_sheet = wb.create_sheet(title=GATTRS)

        # Make header of GLOBALatttributes in first row
        for i, key in enumerate(SHEETS[GATTRS]):
            _ = gattrs_sheet.cell(column=i + 1, row=1, value='{0}'.format(key))

        # Fill sheet with gattrs values
        row_idx = 2
        for key, entries in self.skeleton.gattrs.items():
            for i, entry in enumerate(entries):
                for j, col in enumerate(SHEETS[GATTRS]):
                    if entry[col] is None:
                        entry[col] = ' '
                    _ = gattrs_sheet.cell(column=j + 1, row=row_idx, value='{0}'.format(
                        entry[col]))
                row_idx += 1

        # Create zVariables sheet
        zvars_sheet = wb.create_sheet(title=ZVARS)

        # Make header of zVariables in first row
        for i, key in enumerate(SHEETS[ZVARS]):
            _ = zvars_sheet.cell(column=i + 1, row=1, value='{0}'.format(key))

        # Fill sheet with zvars values
        row_idx = 2
        for zvar, entries in self.skeleton.zvars.items():
            for j, col in enumerate(SHEETS[ZVARS]):
                if entries[col] is None:
                    entry = ' '
                elif isinstance(entries[col], list):
                    entry = ' '.join(entries[col])
                else:
                    entry = entries[col]

                _ = zvars_sheet.cell(column=j + 1, row=row_idx, value='{0}'.format(
                    entry))
            row_idx += 1

        # Create VARIABLEattributes sheet
        vattrs_sheet = wb.create_sheet(title=VATTRS)

        # Make header of VARIABLEattributes in first row
        for i, key in enumerate(SHEETS[VATTRS]):
            _ = vattrs_sheet.cell(column=i + 1, row=1, value='{0}'.format(key))

        # Fill sheet with vattrs values
        row_idx = 2
        for zvar, vattrs in self.skeleton.vattrs.items():
            for vatt, entry in vattrs.items():
                # First column contains Variable name
                _ = vattrs_sheet.cell(column=1, row=row_idx, value='{0}'.format(
                    zvar))
                for j, col in enumerate(SHEETS[VATTRS][1:]):
                    if entry[col] is None:
                        entry[col] = ' '
                    _ = vattrs_sheet.cell(column=j + 2, row=row_idx, value='{0}'.format(
                        entry[col]))
                row_idx += 1

        # Create NRV sheet
        nrv_sheet = wb.create_sheet(title=NRV)

        # Make header of NRV in first row
        for i, key in enumerate(SHEETS[NRV]):
            _ = nrv_sheet.cell(column=i + 1, row=1, value='{0}'.format(key))

        row_idx = 2
        for zvar, fields in self.skeleton.zvars.items():
            if 'NRV' in fields:
                for i, entries in enumerate(fields['NRV']):
                    for j, col in enumerate(SHEETS[NRV]):
                        if entries[col] is None:
                            entry[col] = ' '
                        elif isinstance(entries[col], list):
                            entry = ' '.join(entries[col])
                        else:
                            entry = entries[col]

                        _ = nrv_sheet.cell(column=j + 1, row=row_idx, value='{0}'.format(
                            entry))
                    row_idx += 1

        self.wb = wb
# ________________ Global Functions __________


class Xlsx2skt:
    """Class to convert an Excel 2007 format file into a Skeleton object."""

    def __init__(self, skeleton):

        self.file = None
        self.skeleton = skeleton
        self.cdf_items = dict()

    def parse_xlsx(self, xlsx_file, auto_pad=True):
        """
        Parse the Excel 2007 format file.

        :param: Input Excel 2007 format file
        :return:
        """
        self.file = xlsx_file
        if not os.path.isfile(xlsx_file):
            logger.error('Cannot find Excel file called %s!', xlsx_file)
            raise FileNotFoundError

        if os.path.splitext(xlsx_file)[1] != '.xlsx':
            logger.error('Invalid input Excel format!')
            raise InvalidFile

        logger.info('Parsing %s file...', xlsx_file)
        wkbk = load_workbook(xlsx_file, read_only=True)
        sheet_names = wkbk.sheetnames

        if ('rVariables' in sheet_names) or ('variables' in sheet_names):
            logger.warning('rVariable type is not supported!')

        sheets = dict()
        for shtn in SHEETS:
            logger.debug('Loading %s sheet...', shtn)
            if shtn not in sheet_names:
                logger.error('Missing %s sheet in the input Excel file!', shtn)
                raise InvalidFile
            else:
                wksht = wkbk[shtn]

                sheet_data = OrderedDict()
                for i, row in enumerate(wksht.rows):
                    # Get sheet columns names on the first row
                    # And initialize columns lists
                    if i == 0:
                        header = []
                        for cell in row:
                            if cell.value is None:
                                continue
                            sheet_data[cell.value] = []
                            header.append(cell.value)

                        # Check that file header contains expected columns
                        for col in SHEETS[shtn]:
                            if col not in header:
                                logger.error(
                                    'Missing %s column in the input Excel file!', col)
                                raise InvalidFile

                    # Then, for other rows get cell values for each column
                    else:
                        ncell = len(row)
                        for j, key in enumerate(header):
                            if ncell > j:
                                row_j = str(row[j].value)
                                if row_j.startswith('"') and row_j.endswith('"'):
                                    row_j = row_j[1:-1]
                                else:
                                    row_j = row[j].value
                                sheet_data[key].append(row_j)
                            else:
                                logger.debug('Warning -- empty cell!')
                                sheet_data[key].append(None)

                if not sheet_data:
                    raise InvalidFile('{0} is invalid: '
                                      '{1} sheet has no header nor data!'.format(xlsx_file, shtn))

                sheets[shtn] = sheet_data

        self.cdf_items[GATTRS] = \
            uniq(sheets[GATTRS]['Attribute Name'],
                 not_none=True)
        self.cdf_items[VATTRS] = \
            uniq(sheets[VATTRS]['Attribute Name'],
                 not_none=True)
        self.cdf_items[ZVARS] = \
            uniq(sheets[ZVARS]['Variable Name'],
                 not_none=True)
        logger.debug('%i GLOBAL attributes returned',
                     len(self.cdf_items[GATTRS]))
        logger.debug('%i Variable attributes returned',
                     len(self.cdf_items[VATTRS]))
        logger.debug('%i zVariables returned',
                     len(self.cdf_items[ZVARS]))

        self._xlsx2skt(sheets, auto_pad=auto_pad)

        return self.skeleton

    def _xlsx2skt(self, sheets, auto_pad=True):
        """
        Store input Excel sheets into Skeleton attributes

        :param sheets:
        :return:
        """

        header = dict()
        gattrs = dict()
        vattrs = dict()
        zvars = dict()

        # Fill header
        try:
            for key, value in sheets[HEADER].items():
                header[key] = value[0]
        except:
            logger.exception('{0} sheet is badly formatted, please check!'.format(HEADER))
            raise InvalidFile()

        # Add items number list
        header['nzvar'] = len(self.cdf_items[ZVARS])
        header['ngattr'] = len(self.cdf_items[GATTRS])
        header['nvattr'] = len(self.cdf_items[VATTRS])

        # Fill gattrs
        try:
            for gatt in self.cdf_items[GATTRS]:
                gattrs[gatt] = []
                for i, name in enumerate(sheets[GATTRS]['Attribute Name']):
                    if gatt == name:
                        fields = {}
                        for col in SHEETS[GATTRS]:
                            current_value = sheets[GATTRS][col][i]
                            # If NoneType or empty string, set to one space string
                            if current_value is None or str(current_value).strip() == '':
                                current_value = ' '
                            fields[col] = current_value

                        gattrs[gatt].append(fields)
        except:
            raise InvalidFile(message='{0} sheet is badly formatted, please check!'.format(GATTRS))


        # Fill zvars
        try:
            for i, zvar in enumerate(sheets[ZVARS]['Variable Name']):
                if zvar is None:
                    continue

                fields = dict()
                for col in SHEETS[ZVARS]:
                    try:
                        if sheets[ZVARS][col][i] is None:
                            current_value = ''
                        else:
                            current_value = str(sheets[ZVARS][col][i]).strip()
                        # Fill column values for current zvar
                        if col == 'Sizes':
                            # If Sizes, make sure to have vector
                            fields[col] = [current_value]
                        else:
                            fields[col] = current_value
                    except:
                        raise InvalidEntry(message='Invalid {0} for {1} zVariable!'.format(col, zvar))


                if auto_pad:
                    fields['VAR_PADVALUE'] = assign_pad(
                        sheets[ZVARS]['Data Type'][i])

                zvars[zvar] = fields

                # Add NRV for this zvar (if exist, if not set to empty list [])
                zvars[zvar][NRV] = []
                for i, name in enumerate(sheets[NRV]['Variable Name']):
                    if zvar == name:
                        fields = dict()
                        for col in SHEETS[NRV]:
                            fields[col] = sheets[NRV][col][i]
                        zvars[zvar][NRV].append(fields)

                # Fill variable attributes for this zvar
                for j, name in enumerate(sheets[VATTRS]['Variable Name']):
                    if zvar == name:
                        if zvar not in vattrs:
                            vattrs[zvar] = dict()
                        fields = dict()
                        for col in SHEETS[VATTRS][1:]:
                            fields[col] = sheets[VATTRS][col][j]

                        vattrs[zvar][sheets[VATTRS]['Attribute Name'][j]] = fields
        except:
            logger.exception('{0}, {1} or {2} sheet is badly formatted, '
                             'please check!'.format(ZVARS, VATTRS, NRV))
            raise InvalidFile(message='{0} is invalid!'.format(self.file))

        # Store results into self.skeleton
        self.skeleton.header = header
        self.skeleton.gattrs = gattrs
        self.skeleton.zvars = zvars
        self.skeleton.vattrs = vattrs
        self.skeleton.file = self.file
        self.skeleton.xlsx = True
        self.skeleton.cdf_items = self.cdf_items
        self.skeleton.vattrList = self.cdf_items[VATTRS]

# ________________ Global Functions __________


def assign_pad(data_type):
    """VAR_PADVALUE auto assign.

    Automatically assigns VAR_PADVALUE
    depending to the input data_type
    """
    dtype = data_type.upper()

    if 'EPOCH' in dtype:
        return '01-Jan-0000 00:00:00.000'
    elif 'TT2000' in dtype:
        return '0000-01-01T00:00:00.000000000'
    elif ('INT' in dtype) or ('BYTE' in dtype):
        return '0'
    elif ('FLOAT' in dtype) or ('REAL' in dtype) or ('DOUBLE' in dtype):
        return '0.0'
    elif 'CHAR' in dtype:
        return "\" \""
    else:
        return 'None'

# _________________ Main ____________________________
if __name__ == '__main__':
    print(__file__)
