#!/usr/bin/env python3
# -*- coding: utf-8 -*-

# ________________ HEADER _________________________


"""parse_skt module.

Program to parse an input CDF skeleton table file.

"""


# ________________ IMPORT _________________________
# (Include here the modules to import, e.g. import sys)
import os
import logging
from collections import OrderedDict

from openpyxl import load_workbook

from maser.utils.toolbox import uniq
from maser.utils.cdf.serializer.globals import SHEETS, HEADER, GATTRS, VATTRS, ZVARS, OPTS, NRV
from maser.utils.cdf.serializer.exceptions import InvalidFile
from maser.utils.cdf.serializer.txt import Skt2txt


__all__ = ["Skeleton"]

# ________________ Global Variables _____________
# (define here the global variables)
logger = logging.getLogger(__file__)



# ________________ Class Definition __________
# (If required, define here classes)
class Skeleton():

    def __init__(self,
                 auto_pad=True):
        """
        Init method of the Skeleton class.

        :param skeleton_file:
        """

        self.file = None
        self.header = dict()
        self.gattrs = dict()
        self.vattrs = dict()
        self.zvars = dict()
        self.cdf_items = dict()
        self.xlsx = False
        self.auto_pad = auto_pad

    @staticmethod
    def from_txt(txt_file):
        """
        Create a Skeleton instance from the input text file.

        :param txt_file:
        :return: an instance of Skeleton
        """

        skt = Skeleton()
        skt.parse_txt(txt_file)

        return skt

    @staticmethod
    def from_xlsx(xlsx_file, auto_pad=True):
        """
        Create a Skeleton instance from the input Excel file.

        :param xlsx_file: input Excel file to convert to Skeleton object
        :param auto_pad: If True assigns automatic value to VAR_PADVALUE
        :return: an instance of Skeleton
        """

        skt = Skeleton(auto_pad=auto_pad)
        skt.parse_xlsx(xlsx_file)

        return skt

    def to_txt(self, output_path=None,
               overwrite=False):
        """
        Convert a Skeleton object into an output Skeleton table file in ASCII format.

        :param output_path: Path of the output Skeleton table file
        :param overwrite: If True, the overwrite existing output file
        :return: Path of the output file
        """

        return Skt2txt(self).write_txt(output_path=output_path,
                                              overwrite=overwrite)

#    def to_xlsx(self):

    def parse_xlsx(self, xlsx_file):
        """
        Parse the Excel 2007 format file.

        :param: Input Excel 2007 format file
        :return:
        """
        if not os.path.isfile(xlsx_file):
            logger.error("Cannot find Excel file called %s!", xlsx_file)
            raise FileNotFoundError

        if os.path.splitext(xlsx_file)[1] != ".xlsx":
            logger.error("Invalid input Excel format!")
            raise InvalidFile

        logger.info("Parsing %s file...", xlsx_file)
        wkbk = load_workbook(xlsx_file, read_only=True)
        sheet_names = wkbk.sheetnames

        if ("rVariables" in sheet_names) or ("variables" in sheet_names):
            logger.warning("rVariable type is not supported!")

        sheets = dict()
        for shtn in SHEETS:
            logger.debug("Loading %s sheet...", shtn)
            if shtn not in sheet_names:
                logger.error("Missing %s sheet in the input Excel file!", shtn)
                raise InvalidFile
            else:
                wksht = wkbk[shtn]

                sheet_data = OrderedDict()
                for i, row in enumerate(wksht.rows):
                    # Get sheet columns names on the first row
                    # And initialize columns lists
                    if i == 0:
                        header = []
                        for cell in row:
                            if cell.value is None:
                                continue
                            sheet_data[cell.value] = []
                            header.append(cell.value)

                        # Check that file header contains expected columns
                        for col in SHEETS[shtn]:
                            if col not in header:
                                logger.error("Missing %s column in the input Excel file!", col)
                                raise InvalidFile

                    # Then, get cell values for each column
                    else:
                        ncell = len(row)
                        for j, key in enumerate(header):
                            if ncell > j:
                                row_j = str(row[j].value)
                                if row_j.startswith('"') and row_j.endswith('"'):
                                    row_j = row_j[1:-1]
                                else:
                                    row_j = row[j].value
                                sheet_data[key].append(row_j)
                            else:
                                logger.debug("Warning -- empty cell!")
                                sheet_data[key].append(None)

                sheets[shtn] = sheet_data

        self.cdf_items[GATTRS] = \
            uniq(sheets[GATTRS]["Attribute Name"],
                 not_none=True)
        self.cdf_items[VATTRS] = \
            uniq(sheets[VATTRS]["Attribute Name"],
                 not_none=True)
        self.cdf_items[ZVARS] = \
            uniq(sheets[ZVARS]["Variable Name"],
                 not_none=True)
        logger.debug("%i GLOBAL attributes returned",
                     len(self.cdf_items[GATTRS]))
        logger.debug("%i Variable attributes returned",
                     len(self.cdf_items[VATTRS]))
        logger.debug("%i zVariables returned",
                     len(self.cdf_items[ZVARS]))

        self.file = xlsx_file
        self.xlsx = True
        self._xlsx2skt(sheets)

    def _xlsx2skt(self, sheets):
        """
        Store input Excel sheets into Skeleton attributes

        :param sheets:
        :return:
        """

        # Fill header
        for key, value in sheets[HEADER].items():
            self.header[key] = value[0]

        # Add some options sheet values to "header"
        for opt in ["CDF_COMPRESSION", "CDF_CHECKSUM"]:
            self.header[opt] = sheets[OPTS][opt][0]

        # Add items number list
        self.header["nzvar"] = len(self.cdf_items[ZVARS])
        self.header["ngattr"] = len(self.cdf_items[GATTRS])
        self.header["nvattr"] = len(self.cdf_items[VATTRS])

        # Fill gattrs
        for gatt in self.cdf_items[GATTRS]:
            self.gattrs[gatt] = []
            for i, name in enumerate(sheets[GATTRS]["Attribute Name"]):
                if gatt == name:
                    fields = {}
                    for col in SHEETS[GATTRS][1:]:
                        fields[col] = sheets[GATTRS][col][i]

                    self.gattrs[gatt].append(fields)

        # Fill zvars
        for i, zvar in enumerate(sheets[ZVARS]["Variable Name"]):
            if zvar is None:
                continue

            fields = dict()
            for col in SHEETS[ZVARS][1:]:
                fields[col] = sheets[ZVARS][col][i]
            # Add option
            for opt in ["VAR_COMPRESSION", "VAR_SPARESERECORDS"]:
                fields[opt] = sheets[OPTS][opt][0]

            if self.auto_pad:
                fields["VAR_PADVALUE"] = assign_pad(sheets[ZVARS]["Data Type"][i])
            else:
                fields["VAR_PADVALUE"] = sheets[OPTS]["VAR_PADVALUE"][0]
            self.zvars[zvar] = fields

            # Add NRV for this zvar (if exist, if not set to "None")
            self.zvars[zvar][NRV] = []
            for i, name in enumerate(sheets[NRV]["Variable Name"]):
                if zvar == name:
                    fields = dict()
                    for col in SHEETS[NRV][1:]:
                        fields[col] = sheets[NRV][col][i]
                    self.zvars[zvar][NRV].append(fields)

            # Fill variable attributes for this zvar
            for j, name in enumerate(sheets[VATTRS]["Variable Name"]):
                if zvar == name:
                    if zvar not in self.vattrs:
                        self.vattrs[zvar] = dict()
                    fields = dict()
                    for col in SHEETS[VATTRS][2:]:
                        fields[col] = sheets[VATTRS][col][j]

                    self.vattrs[zvar][sheets[VATTRS]["Attribute Name"][j]] = fields

    def parse_txt(self, txt_file):
        """
        Method to parse an input CDF skeleton table file.
        (rVariable parsing not supported)

        :param skeleton:
        :return: instance of a Skeleton object
        """

        if not os.path.isfile(txt_file):
            logger.error("{0} NOT FOUND, ABORTING!".format(txt_file))
            raise FileNotFoundError

        with open(txt_file, 'rt') as skt_file:
            for row in skt_file.readlines():

                # Read header
                if row.startswith(HEADER):
                    self.section = HEADER

                # Read GLOBALattributes
                elif row.startswith(GATTRS):
                    self.section = GATTRS

                # Read VARIABLEattributes
                elif row.startswith(VATTRS):
                    self.section = VATTRS

                # Read zVariables
                elif row.startswith(ZVARS):
                    self.section = ZVARS

                self._extract(row.strip())

# ________________ Global Functions __________
def assign_pad(data_type):
    """VAR_PADVALUE auto assign.

    Automatically assigns VAR_PADVALUE
    depending to the input data_type
    """
    dtype = data_type.upper()

    if "EPOCH" in dtype:
        return "01-Jan-0000 00:00:00.000"
    elif "TT2000" in dtype:
        return "0000-01-01T00:00:00.000000000"
    elif ("INT" in dtype) or ("BYTE" in dtype):
        return "0"
    elif ("FLOAT" in dtype) or ("REAL" in dtype) or ("DOUBLE" in dtype):
        return "0.0"
    elif "CHAR" in dtype:
        return "\" \""
    else:
        return "None"


# _________________ Main ____________________________
if __name__ == "__main__":
    print(__file__)