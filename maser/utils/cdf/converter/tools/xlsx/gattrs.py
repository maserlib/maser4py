#! /usr/bin/env python
# -*- coding: utf-8 -*-

"""Python module cdf.converter.tools.xlsx.gattr."""

# ________________ IMPORT _________________________
# (Include here the modules to import, e.g. import sys)
import os.path as asp
import logging
from glob import glob

from openpyxl import load_workbook as lwb

from .gen import add_row

__all__ = ["add_gattr",
            "set_gattr_entries",
            "set_gattr_dtype",
            "add_gattr_entry",
            "rm_gattr"]

# ________________ HEADER _________________________

# Mandatory
__version__ = ""
__author__ = "Xavier Bonnin"
__date__ = "2016-09-30"

# Optional
__license__ = ""
__credit__ = [""]
__maintainer__ = ""
__email__ = ""
__project__ = "MASER"
__institute__ = "LESIA, Observatoire de Paris, LESIA"
__changes__ = ""


# ________________ Global Variables _____________
# (define here the global variables)
logger = logging.getLogger(__name__)

GATT_SHEET_NAME = "GLOBALattributes"


# ________________ Class Definition __________
# (If required, define here classes)
class GattrException(Exception):
    pass

# ________________ Global Functions __________
# (If required, define here gobal functions)


def check_args(func):
    """Decorator for arg checking."""
    def wrapper(*args, **kwargs):

        xlsx = args[0]
        attname = args[1]

        if not asp.isfile(xlsx):
            logger.error("Input file not found [{0}]!".format(xlsx))
            raise GattrException
        else:
            wb = lwb(xlsx)

        shnames = wb.get_sheet_names()
        if GATT_SHEET_NAME not in shnames:
            logger.error("Input file does not contain {0} sheet!".format(
                                        GATT_SHEET_NAME))
            raise GattrException

        if "overwrite" not in kwargs:
            kwargs["overwrite"] = False
        overwrite = kwargs["overwrite"]

        if "output" not in kwargs or kwargs["output"] is None:
            kwargs["output"] = xlsx
        output = kwargs["output"]

        wb = func(*args, **kwargs)
        if wb is None:
            return False

        if overwrite is False:
            nfile = len(glob(output + "*"))
            if nfile > 0:
                output = output + "." + str(nfile)
        logger.info("{0} attribute added to {1}".format(attname, output))
        wb.save(output)

        return True

    return wrapper


@check_args
def add_gattr(xlsx, attname, cdftype, entries,
              output=None,
              overwrite=False):
    """Add a global attribute into the input CDF Excel skeleton."""
    if type(entries) is not list:
        entries = list(entries)

    wb = lwb(xlsx)

    # Get sheet for global attributes
    ws = wb.get_sheet_by_name(GATT_SHEET_NAME)

    # Check if the gattr alreay exists
    max_row = ws.max_row + 1
    cells = [ws["A{0}".format(i)].value
        for i in range(1, max_row)
        if ws["A{0}".format(i)].value == attname]

    if len(cells) > 0:
        logger.warning("{0} attribute already exists!".format(attname))
        return None

    # Add the new gattr entries at the bottom of the
    # sheet
    for i, entry in enumerate(entries):
        index = max_row + i

        # print(i, entry, index)

        # Add gatt name
        ws.cell(row=index, column=1).value = str(attname)
        # Add entry number
        ws.cell(row=index, column=2).value = i + 1
        # Add cdf dtype
        ws.cell(row=index, column=3).value = str(cdftype)
        # Add entry value
        ws.cell(row=index, column=4).value = entry

    return wb


@check_args
def set_gattr_entries(xlsx, attname, new_entries,
                      output=None, overwrite=False):
    """set_gattr_entries.

    @author: X.Bonnin, LESIA, Obs. Paris, CNRS

    Update global attribute entries into the input CDF Excel skeleton.

    Positional arguments:
        xlsx - CDF Excel skeleton file
        attname - Name of the global attribute
        new_entries - dictionary containing the
                      entry indexes as keys and
                      entry values as values
                      (e.g. {"1":"value1", "2":value2})

    Optional keywords:
        output - Name of the output file
        overwrite - If true, then overwrite existing output file

    """
    wb = lwb(xlsx)

    if type(new_entries) is not dict:
        logger.error("new_entries must a dictionary!")
        raise GattrException

    # Get sheet for global attributes
    ws = wb.get_sheet_by_name(GATT_SHEET_NAME)

    # Get indices of row for the gattr
    max_row = ws.max_row + 1
    rows = [i
        for i in range(1, max_row)
        if ws["A{0}".format(i)].value == attname]

    if len(rows) == 0:
        logger.warning("There is no {0} attribute in {1}!".format(
                                        attname, xlsx))
        return None

    ninserted = 0
    for i, row in enumerate(rows):
        entry_num = str(ws["B{0}".format(row)].value)
        if entry_num in new_entries:
            ws["D{0}".format(row)].value = new_entries[entry_num]
            ninserted += 1
        else:
            logger.warning("No change brung to entry #{0}".format(
                                                    entry_num))

    if ninserted != len(new_entries):
        logger.warning("attribute entries have not been updated!")

    return wb


@check_args
def set_gattr_dtype(xlsx, attname, new_dtype,
                    output=None, overwrite=False):
    """set_gattr_dtype.

    Update the CDF data type of a given global attribute
    """
    wb = lwb(xlsx)

    # Get sheet for global attributes
    ws = wb.get_sheet_by_name(GATT_SHEET_NAME)

    # Get indices of row for the gattr
    max_row = ws.max_row + 1
    rows = [i
        for i in range(1, max_row)
        if ws["A{0}".format(i)].value == attname]

    if len(rows) == 0:
        logger.warning("There is no {0} attribute in {1}!".format(
                                        attname, xlsx))
        return None

    # Update CDF dtype
    for row in rows:
        ws["C{0}".format(row)].value = new_dtype

    return wb


@check_args
def add_gattr_entry(xlsx, attname, value,
            output=None, overwrite=False):
    """add_gattr_entry.

    Add an entry to an existing attribute "attname"
    in the xlsx file.
    """
    wb = lwb(xlsx)

    # Get sheet for global attributes
    ws = wb.get_sheet_by_name(GATT_SHEET_NAME)

    # Get indices of row for the gattr
    max_row = ws.max_row + 1
    rows = [i
        for i in range(1, max_row)
        if ws["A{0}".format(i)].value == attname]

    if len(rows) == 0:
        logger.warning("There is no {0} attribute in {1}!".format(
                                        attname, xlsx))
        return None

    entry_num = len(rows) + 1
    values = [attname, entry_num, ws["C{0}".format(rows[0])].value, value]
    wb = add_row(xlsx, GATT_SHEET_NAME, rows[-1] + 1, values=values)

    return wb


@check_args
def rm_gattr(xlsx, attname,
             output=None, overwrite=False):
    """rm_gattr.

    Remove a global attr. from the input CDF Excel skeleton file.
    (By default a new file is created.)

    """
    wb = lwb(xlsx)

    # Get sheet for global attributes
    ws = wb.get_sheet_by_name(GATT_SHEET_NAME)

    # Get indices of row for the gattr
    max_row = ws.max_row + 1
    rows = [i
        for i in range(1, max_row)
        if ws["A{0}".format(i)].value == attname]

    if len(rows) == 0:
        logger.warning("There is no {0} attribute in {1}!".format(
                                        attname, xlsx))
        return None

    # Update CDF dtype
    for row in rows:
        # Add gatt name
        ws.cell(row=row, column=1).value = None
        # Add entry number
        ws.cell(row=row, column=2).value = None
        # Add cdf dtype
        ws.cell(row=row, column=3).value = None
        # Add entry value
        ws.cell(row=row, column=4).value = None

    return wb


def main():
    """Main program."""
    logger.info("This is the cdf.converter.tools.xlsx.gattrs module.")

# _________________ Main ____________________________
if (__name__ == "__main__"):
    main()
