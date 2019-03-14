#! /usr/bin/env python
# -*- coding: utf-8 -*-

"""Python module cdf.converter.tools.xlsx.vattrs."""

# ________________ IMPORT _________________________
# (Include here the modules to import, e.g. import sys)
import logging


from openpyxl import load_workbook as lwb

from .gen import add_row, get_row, get_item_list, check_args

__all__ = ["add_vattr",
            "set_vattr_entries",
            "set_vattr_dtype",
            "set_vattr_var",
            "rm_vattr",
            "rm_var",
            "set_vattr_var"]

# ________________ HEADER _________________________



# ________________ Global Variables _____________
# (define here the global variables)
logger = logging.getLogger(__name__)

VATT_SHEET_NAME = "VARIABLEattributes"


# ________________ Class Definition __________
# (If required, define here classes)
class VattrException(Exception):
    pass

# ________________ Global Functions __________
# (If required, define here gobal functions)


@check_args(VATT_SHEET_NAME)
def add_var(xlsx, varname, entries,
            insert_row=None,
            output=None,
            overwrite=False):
    """
    Add_var method.

    Add a variable and its vattribute(s)
    into the input CDF Excel skeleton.

    varname is a string scalar containing the name of the CDF variable
    entries is a list of n sub-lists containing the following items:
        [vattribute name, data type, value]
    There must be one sub-list per vattribute to insert.

    """
    wb = lwb(xlsx)

    # Get sheet for variable attributes
    ws = wb[VATT_SHEET_NAME]

    # Check if the variable exists
    rows = get_row(ws, varname, column="A")

    if len(rows) > 0:
        logger.warning(
                       "{0} variable already exists!".format(
                                                    varname))
        return wb

    # By default, insert vattributes at the end of the sheet
    if insert_row is None:
        insert_row = ws.max_row + 1

    for i, entry in enumerate(entries):
        if len(entry) != 3:
            logger.warning(
                "Wrong number of elements for entry #{0} [{1}]".format(
                                                i, ", ".join(entry)))
            continue
        values = [varname, entry[0], entry[1], entry[2]]
        wb = add_row(xlsx, VATT_SHEET_NAME, insert_row + i,
                     values=values,
                     output=output, overwrite=True)

    return wb


@check_args(VATT_SHEET_NAME)
def add_vattr(xlsx, attname, cdftype, entry,
              varname=None,
              output=None,
              overwrite=False):
    """Add a variable attribute into the input CDF Excel skeleton."""
    wb = lwb(xlsx)

    # Get sheet for variable attributes
    ws = wb[VATT_SHEET_NAME]

    # Check if the variable attribute exists
    rows = get_row(ws, attname, column="B")

    if len(rows) > 0:
        logger.warning(
                       "{0} variable attribute already exists!".format(
                                                    attname))
        return wb

    # Get list of variables
    zvars = get_item_list(ws)

    # Add the variable attribute for each var
    # (except if the varname optional argument is set)
    for zvar in zvars:
        if varname is not None and zvar != varname:
            continue
        rows = get_row(ws, zvar)
        values = [zvar, attname, cdftype, entry]
        wb = add_row(xlsx, VATT_SHEET_NAME, rows[-1] + 1, values=values,
                     output=output, overwrite=True)

    return wb


@check_args(VATT_SHEET_NAME)
def set_vattr_entries(xlsx, attname, new_entry,
                      varname=None,
                      output=None, overwrite=False):
    """set_vattr_entries.

    @author: X.Bonnin, LESIA, Obs. Paris, CNRS

    Update variable attribute entry into the input CDF Excel skeleton.

    Positional arguments:
        xlsx - CDF Excel skeleton file
        attname - Name of the variable attribute
        new_entry - a scalar containing the new entry

    Optional keywords:
        varname - if a variable name is provided, then only
                  set the vattribute for this variable
        output - Name of the output file
        overwrite - If true, then overwrite existing output file

    """
    wb = lwb(xlsx)

    # Get sheet for variable attributes
    ws = wb[VATT_SHEET_NAME]

    # Get indices of row for the vattr
    rows = get_row(ws, attname, column="B")

    if len(rows) == 0:
        logger.warning("There is no {0} attribute!".format(
                                        attname))
        return None

    ninserted = 0
    for row in rows:
        zvar = ws["A{0}".format(row)].value
        if varname is not None and varname != zvar:
            continue
        ws["D{0}".format(row)].value = new_entry
        ninserted += 1

    if ninserted == 0:
        logger.warning("{0} attribute has not been updated!".format(attname))

    return wb


@check_args(VATT_SHEET_NAME)
def set_vattr_dtype(xlsx, attname, new_dtype,
                    varname=None,
                    output=None, overwrite=False):
    """set_vattr_dtype.

    Update the CDF data type of a given variable attribute
    """
    wb = lwb(xlsx)

    # Get sheet for variable attributes
    ws = wb[VATT_SHEET_NAME]

    # Get indices of row for the vattr
    rows = get_row(ws, attname, column="B")

    if len(rows) == 0:
        logger.warning("There is no {0} attribute!".format(
                                        attname))
        return None

    # Update CDF dtype
    for row in rows:
        zvar = ws["A{0}".format(row)].value
        if varname is not None and varname != zvar:
            continue
        ws["C{0}".format(row)].value = new_dtype

    return wb


@check_args(VATT_SHEET_NAME)
def rename_vattr(xlsx, old_attname, new_attname,
                 varname=None,
                 output=None, overwrite=False):
    """rename_vattr.

    Rename a variable attr. from the input CDF Excel skeleton file.
    (By default a new file is created.)

    """
    wb = lwb(xlsx)

    # Get sheet for variable attributes
    ws = wb[VATT_SHEET_NAME]

    # Get indices of row for the vattr
    rows = get_row(ws, old_attname, column="B")

    if len(rows) == 0:
        logger.warning("There is no {0} attribute!".format(
                                        old_attname))
        return wb

    for row in rows:
        zvar = ws["A{0}".format(row)].value
        if varname is not None and varname != zvar:
            continue
        # Change vatt name
        ws.cell(row=row, column=2).value = new_attname

    return wb


@check_args(VATT_SHEET_NAME)
def rm_vattr(xlsx, attname,
             varname=None,
             output=None, overwrite=False):
    """rm_vattr.

    Remove a variable attr. from the input CDF Excel skeleton file.
    (By default a new file is created.)

    """
    wb = lwb(xlsx)

    # Get sheet for variable attributes
    ws = wb[VATT_SHEET_NAME]

    # Get indices of row for the vattr
    rows = get_row(ws, attname, column="B")

    if len(rows) == 0:
        logger.warning("There is no {0} attribute!".format(
                                        attname))
        return wb

    # Update Cells to put tuple of None for the given vattr
    for row in rows:
        zvar = ws["A{0}".format(row)].value
        if varname is not None and varname != zvar:
            continue
        # Remove var name
        ws.cell(row=row, column=1).value = None
        # Remove attribute name
        ws.cell(row=row, column=2).value = None
        # Remove cdf dtype
        ws.cell(row=row, column=3).value = None
        # Remove entry value
        ws.cell(row=row, column=4).value = None

    return wb


@check_args(VATT_SHEET_NAME)
def rm_var(xlsx, varname,
             output=None, overwrite=False):
    """rm_var.

    Remove a variable in the vattrs sheet
    from the input CDF Excel skeleton file.
    (By default a new file is created.)

    """
    wb = lwb(xlsx)

    # Get sheet for variable attributes
    ws = wb[VATT_SHEET_NAME]

    # Get indices of row for the vattr
    rows = get_row(ws, varname, column="A")

    if len(rows) == 0:
        logger.warning("There is no {0} variable!".format(
                                        varname))
        return wb

    # Update Cells to put tuple of None for the given variable
    for row in rows:
        # Remove var name
        ws.cell(row=row, column=1).value = None
        # Remove attribute name
        ws.cell(row=row, column=2).value = None
        # Remove cdf dtype
        ws.cell(row=row, column=3).value = None
        # Remove entry value
        ws.cell(row=row, column=4).value = None

    return wb


@check_args(VATT_SHEET_NAME)
def set_vattr_var(xlsx, old_varname, new_varname,
                  output=None, overwrite=False):
    """Rename a variable name in the variable attribute sheet."""
    wb = lwb(xlsx)

    # Get sheet for variable attributes
    ws = wb[VATT_SHEET_NAME]

    # Check if the variable attribute exists
    rows = get_row(ws, old_varname)

    if len(rows) == 0:
        logger.warning(
                       "{0} variable does not exist!".format(
                                                    old_varname))
        return wb

    for row in rows:
        ws.cell(row=row, column=1).value = new_varname

    return wb


def main():
    """Main program."""
    logger.info("This is the cdf.converter.tools.xlsx.vattrs module.")

# _________________ Main ____________________________
if (__name__ == "__main__"):
    main()
