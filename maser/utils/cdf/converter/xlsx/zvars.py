#! /usr/bin/env python
# -*- coding: utf-8 -*-

"""Python module cdf.converter.tools.xlsx.zvars."""

# ________________ IMPORT _________________________
# (Include here the modules to import, e.g. import sys)
import logging


from openpyxl import load_workbook as lwb

from .gen import add_row, get_row, check_args

__all__ = ["add_zvar",
            "set_zvar_entry",
            "rm_zvar",
            "rename_zvar"]

# ________________ HEADER _________________________



# ________________ Global Variables _____________
# (define here the global variables)
logger = logging.getLogger(__name__)

ZVAR_SHEET_NAME = "zVariables"


# ________________ Class Definition __________
# (If required, define here classes)
class zVarException(Exception):
    pass

# ________________ Global Functions __________
# (If required, define here gobal functions)


@check_args(ZVAR_SHEET_NAME)
def add_zvar(xlsx, entry,
              insert_row=None,
              output=None,
              overwrite=False):
    """Add a zVariable into the input CDF Excel skeleton.

    entry = [variable name, data type, number elements,
                dims, sizes, record variance, Dimension variances]

    or if dims=0

    entry = [variable name, data type, number elements,
                dims, sizes, record variance]


    None value in the entry list will be let blank
    """
    wb = lwb(xlsx)

    # Get sheet for zVars
    ws = wb[ZVAR_SHEET_NAME]

    varname = entry[0]
    # Check if the zVariable already exists
    rows = get_row(ws, varname, column="A")

    if len(rows) > 0:
        logger.warning(
                       "{0} variable already exists!".format(
                                                    varname))
        return wb

    # Add the zvariable
    if len(entry) == 7:
        values = entry
    elif len(entry) == 6:
        values = entry.append(None)
    else:
        msg = "Wrong number of elements in entry!"
        logger.error(msg)
        raise zVarException(msg)

    if insert_row is None:
        insert_row = ws.max_row + 1

    wb = add_row(xlsx, ZVAR_SHEET_NAME, insert_row + 1,
                 values=values, output=output, overwrite=True)

    return wb


@check_args(ZVAR_SHEET_NAME)
def set_zvar_entry(xlsx, varname,
                    dtype=None,
                    NumElems=None,
                    Dims=None,
                    Sizes=None,
                    RecVar=None, DimVars=None,
                    output=None, overwrite=False):
    """set_zvar_entries.

    @author: X.Bonnin, LESIA, Obs. Paris, CNRS

    Update zvariable entry into the input CDF Excel skeleton.

    """
    wb = lwb(xlsx)

    # Get sheet for variable attributes
    ws = wb[ZVAR_SHEET_NAME]

    # Get indices of row for the zvar
    row = get_row(ws, varname, column="A")

    if len(row) == 0:
        logger.warning("There is no {0} zVariable!".format(
                                        varname))
        return None

    ninserted = 0
    if dtype is not None:
        ws["B{0}".format(row)].value = dtype
        ninserted += 1
    if NumElems is not None:
        ws["C{0}".format(row)].value = NumElems
        ninserted += 1
    if Dims is not None:
        ws["D{0}".format(row)].value = Dims
        ninserted += 1
    if Sizes is not None:
        ws["E{0}".format(row)].value = Sizes
        ninserted += 1
    if RecVar is not None:
        ws["F{0}".format(row)].value = RecVar
        ninserted += 1
    if DimVars is not None:
        ws["G{0}".format(row)].value = DimVars
        ninserted += 1

    if ninserted == 0:
        logger.warning("{0} variable has not been updated!".format(varname))

    return wb


@check_args(ZVAR_SHEET_NAME)
def rename_zvar(xlsx, old_varname, new_varname,
                 output=None, overwrite=False):
    """rename_zvar.

    Rename a variable from the input CDF Excel skeleton file.
    (By default a new file is created.)

    """
    wb = lwb(xlsx)

    # Get sheet for zVariables
    ws = wb[ZVAR_SHEET_NAME]

    # Get indices of row for the zvar
    rows = get_row(ws, old_varname)

    if len(rows) == 0:
        logger.warning("There is no {0} variable!".format(
                                        old_varname))
        return None

    for row in rows:
        # Change gatt name
        ws.cell(row=row, column=1).value = new_varname

    return wb


@check_args(ZVAR_SHEET_NAME)
def rm_zvar(xlsx, varname,
             output=None, overwrite=False):
    """rm_zvar.

    Remove a variable from the input CDF Excel skeleton file.
    (By default a new file is created.)

    """
    wb = lwb(xlsx)

    # Get sheet for zVariables
    ws = wb[ZVAR_SHEET_NAME]

    # Get indices of row for the zvar
    rows = get_row(ws, varname)

    if len(rows) == 0:
        logger.warning("There is no {0} variable!".format(
                                        varname))
        return None

    # Update Cells
    for row in rows:
        # Remove var name
        ws.cell(row=row, column=1).value = None
        # Remove entry number
        ws.cell(row=row, column=2).value = None
        # Remove cdf dtype
        ws.cell(row=row, column=3).value = None
        # Remove Dims
        ws.cell(row=row, column=4).value = None
        # Remove Sizes
        ws.cell(row=row, column=5).value = None
        # Remove Record Variance
        ws.cell(row=row, column=6).value = None
        # Remove Dims Variance
        ws.cell(row=row, column=7).value = None

    return wb


def main():
    """Main program."""
    logger.info("This is the cdf.converter.tools.xlsx.zvars module.")

# _________________ Main ____________________________
if (__name__ == "__main__"):
    main()
