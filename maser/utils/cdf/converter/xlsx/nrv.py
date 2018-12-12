#! /usr/bin/env python
# -*- coding: utf-8 -*-

"""Python module cdf.converter.tools.xlsx.nrv."""

# ________________ IMPORT _________________________
# (Include here the modules to import, e.g. import sys)
import logging

from openpyxl import load_workbook as lwb

from .gen import add_row, get_row, check_args

__all__ = ["add_nrv",
            "set_nrv_entries",
            "add_nrv_entry",
            "rm_nrv"]

# ________________ HEADER _________________________


# ________________ Global Variables _____________
# (define here the global variables)
logger = logging.getLogger(__name__)

NRV_SHEET_NAME = "NRV"


# ________________ Class Definition __________
# (If required, define here classes)
class NrvException(Exception):
    pass

# ________________ Global Functions __________
# (If required, define here gobal functions)


@check_args(NRV_SHEET_NAME)
def add_nrv(xlsx, varname, entries,
              output=None,
              overwrite=False):
    """Add a global attribute into the input CDF Excel skeleton."""
    if type(entries) is not list:
        entries = list(entries)

    wb = lwb(xlsx)

    # Get sheet for global attributes
    ws = wb[NRV_SHEET_NAME]

    # Check if the NRV variable alreay exists
    max_row = ws.max_row + 1
    cells = get_row(ws, varname)

    if len(cells) > 0:
        logger.warning("{0} NRV variable already exists!".format(varname))
        return ws

    # Add the new NRV variable entries at the bottom of the
    # sheet
    for i, entry in enumerate(entries):
        index = max_row + i

        # print(i, entry, index)

        # Add NRV var name
        ws.cell(row=index, column=1).value = str(varname)
        # Add index number
        ws.cell(row=index, column=2).value = i + 1
        # Add entry value
        ws.cell(row=index, column=3).value = entry

    return wb


@check_args(NRV_SHEET_NAME)
def set_nrv_entries(xlsx, varname, new_entries,
                      output=None, overwrite=False):
    """set_nrv_entries.

    @author: X.Bonnin, LESIA, Obs. Paris, CNRS

    Update NRV entries into the input CDF Excel skeleton.

    Positional arguments:
        xlsx - CDF Excel skeleton file
        varname - Name of the CDF variable to set in the NRV sheet
        new_entries - dictionary containing the
                      entry indexes as keys and
                      entry values as values
                      (e.g. {"1":"value1", "2":value2})

    Optional keywords:
        output - Name of the output file
        overwrite - If true, then overwrite existing output file

    """
    wb = lwb(xlsx)

    if type(new_entries) is not dict:
        logger.error("NEW_ENTRIES input must a dictionary!")
        raise NrvException

    # Get sheet for global attributes
    ws = wb[NRV_SHEET_NAME]

    # Get indices of row for the var
    rows = get_row(ws, varname)

    if len(rows) == 0:
        logger.warning("There is no {0} NRV variable!".format(
                                        varname))
        return None

    ninserted = 0
    for i, row in enumerate(rows):
        entry_num = str(ws["B{0}".format(row)].value)
        if entry_num in new_entries:
            ws["C{0}".format(row)].value = new_entries[entry_num]
            ninserted += 1
        # else:
        #    logger.warning("No change brung to entry #{0}".format(
        #                                            entry_num))

    if ninserted != len(new_entries):
        logger.warning("NRV variable entries have not been updated!")

    return wb


@check_args(NRV_SHEET_NAME)
def add_nrv_entry(xlsx, varname, value,
            output=None, overwrite=False):
    """add_nrv_entry.

    Add an entry to an existing variable "attname"
    in the NRV sheet of the input xlsx file.
    """
    wb = lwb(xlsx)

    # Get sheet for global attributes
    ws = wb[NRV_SHEET_NAME]

    # Get indices of row for the NRV variable
    rows = get_row(ws, varname)

    if len(rows) == 0:
        logger.warning("There is no {0} NRV variable!".format(
                                        varname))
        return None

    entry_num = len(rows) + 1
    values = [varname, entry_num, value]
    wb = add_row(xlsx, NRV_SHEET_NAME, rows[-1] + 1, values=values)

    return wb


@check_args(NRV_SHEET_NAME)
def rename_nrv(xlsx, old_varname, new_varname,
                 output=None, overwrite=False):
    """rename_nrv.

    Rename a NRV variable from the input CDF Excel skeleton file.
    (By default a new file is created.)

    """
    wb = lwb(xlsx)

    # Get sheet for global attributes
    ws = wb[NRV_SHEET_NAME]

    # Get indices of row for the NRV variable
    rows = get_row(ws, old_varname)

    if len(rows) == 0:
        logger.warning("There is no {0} NRV variable!".format(
                                        old_varname))
        return None

    for row in rows:
        # Change var name
        ws.cell(row=row, column=1).value = new_varname

    return wb


@check_args(NRV_SHEET_NAME)
def rm_nrv(xlsx, varname,
             output=None, overwrite=False):
    """rm_nrv method.

    Remove the entries of a given NRV variable
    from the input CDF Excel skeleton file.
    (By default a new file is created.)

    """
    wb = lwb(xlsx)

    # Get sheet for global attributes
    ws = wb[NRV_SHEET_NAME]

    # Get indices of row for the NRV variable
    rows = get_row(ws, varname)

    if len(rows) == 0:
        logger.warning("There is no {0} NRV variable!".format(
                                        varname))
        return None

    # Remove entries
    for row in rows:
        # Remove var name
        ws.cell(row=row, column=1).value = None
        # Remove entry number
        ws.cell(row=row, column=2).value = None
        # Remove entry value
        ws.cell(row=row, column=3).value = None

    return wb


def main():
    """Main program."""
    logger.info("This is the cdf.converter.tools.xlsx.nrv module.")

# _________________ Main ____________________________
if (__name__ == "__main__"):
    main()
