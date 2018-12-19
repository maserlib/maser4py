#! /usr/bin/env python
# -*- coding: utf-8 -*-

"""Python module cdf.converter.tools.xlsx.gattr."""

# ________________ IMPORT _________________________
# (Include here the modules to import, e.g. import sys)
import logging

from openpyxl import load_workbook as lwb

from .gen import add_row, get_row, check_args

__all__ = ["add_gattr",
            "set_gattr_entries",
            "set_gattr_dtype",
            "add_gattr_entry",
            "rm_gattr"]

# ________________ HEADER _________________________



# ________________ Global Variables _____________
# (define here the global variables)
logger = logging.getLogger(__name__)

GATT_SHEET_NAME = "GLOBALattributes"


# ________________ Class Definition __________
# (If required, define here classes)
class GattrException(Exception):
    pass

# ________________ Global Functions __________
# (If required, define here gobal functions)


@check_args(GATT_SHEET_NAME)
def add_gattr(xlsx, attname, cdftype, entries,
              output=None,
              overwrite=False):
    """Add a global attribute into the input CDF Excel skeleton."""
    if type(entries) is not list:
        entries = list(entries)

    wb = lwb(xlsx)

    # Get sheet for global attributes
    ws = wb[GATT_SHEET_NAME]

    # Check if the gattr alreay exists
    max_row = ws.max_row + 1
    cells = get_row(ws, attname)

    if len(cells) > 0:
        logger.warning("{0} attribute already exists!".format(attname))
        return wb

    # Add the new gattr entries at the bottom of the
    # sheet
    for i, entry in enumerate(entries):
        index = max_row + i

        # print(i, entry, index)

        # Add gatt name
        ws.cell(row=index, column=1).value = str(attname)
        # Add entry number
        ws.cell(row=index, column=2).value = i + 1
        # Add cdf dtype
        ws.cell(row=index, column=3).value = str(cdftype)
        # Add entry value
        ws.cell(row=index, column=4).value = entry

    return wb


@check_args(GATT_SHEET_NAME)
def set_gattr_entries(xlsx, attname, new_entries,
                      output=None, overwrite=False):
    """set_gattr_entries.

    @author: X.Bonnin, LESIA, Obs. Paris, CNRS

    Update global attribute entries into the input CDF Excel skeleton.

    Positional arguments:
        xlsx - CDF Excel skeleton file
        attname - Name of the global attribute
        new_entries - dictionary containing the
                      entry indexes as keys and
                      entry values as values
                      (e.g. {"1":"value1", "2":value2})

    Optional keywords:
        output - Name of the output file
        overwrite - If true, then overwrite existing output file

    """
    wb = lwb(xlsx)

    if type(new_entries) is not dict:
        logger.error("new_entries must a dictionary!")
        raise GattrException

    # Get sheet for global attributes
    ws = wb[GATT_SHEET_NAME]

    # Get indices of row for the gattr
    rows = get_row(ws, attname)

    if len(rows) == 0:
        logger.warning("There is no {0} attribute!".format(
                                        attname))
        return None

    ninserted = 0
    for i, row in enumerate(rows):
        entry_num = str(ws["B{0}".format(row)].value)
        if entry_num in new_entries:
            ws["D{0}".format(row)].value = new_entries[entry_num]
            ninserted += 1
        # else:
        #    logger.warning("No change brung to entry #{0}".format(
        #                                            entry_num))

    if ninserted != len(new_entries):
        logger.warning("Attribute entries have not been updated!")

    return wb


@check_args(GATT_SHEET_NAME)
def set_gattr_dtype(xlsx, attname, new_dtype,
                    output=None, overwrite=False):
    """set_gattr_dtype.

    Update the CDF data type of a given global attribute
    """
    wb = lwb(xlsx)

    # Get sheet for global attributes
    ws = wb[GATT_SHEET_NAME]

    # Get indices of row for the gattr
    rows = get_row(ws, attname)

    if len(rows) == 0:
        logger.warning("There is no {0} attribute!".format(
                                        attname))
        return None

    # Update CDF dtype
    for row in rows:
        ws["C{0}".format(row)].value = new_dtype

    return wb


@check_args(GATT_SHEET_NAME)
def add_gattr_entry(xlsx, attname, value,
            output=None, overwrite=False):
    """add_gattr_entry.

    Add an entry to an existing attribute "attname"
    in the xlsx file.
    """
    wb = lwb(xlsx)

    # Get sheet for global attributes
    ws = wb[GATT_SHEET_NAME]

    # Get indices of row for the gattr
    rows = get_row(ws, attname)

    if len(rows) == 0:
        logger.warning("There is no {0} attribute!".format(
                                        attname))
        return None

    entry_num = len(rows) + 1
    # Values = [attribute name, entry index number, cdf data type, value]
    values = [attname, entry_num, ws["C{0}".format(rows[0])].value, value]
    wb = add_row(xlsx, GATT_SHEET_NAME, rows[-1] + 1, values=values)

    return wb


@check_args(GATT_SHEET_NAME)
def rename_gattr(xlsx, old_attname, new_attname,
                 output=None, overwrite=False):
    """rename_gattr.

    Rename a global attr. from the input CDF Excel skeleton file.
    (By default a new file is created.)

    """
    wb = lwb(xlsx)

    # Get sheet for global attributes
    ws = wb[GATT_SHEET_NAME]

    # Get indices of row for the gattr
    rows = get_row(ws, old_attname)

    if len(rows) == 0:
        logger.warning("There is no {0} attribute!".format(
                                        old_attname))
        return None

    for row in rows:
        # Change gatt name
        ws.cell(row=row, column=1).value = new_attname

    return wb


@check_args(GATT_SHEET_NAME)
def rm_gattr(xlsx, attname,
             output=None, overwrite=False):
    """rm_gattr.

    Remove a global attr. from the input CDF Excel skeleton file.
    (By default a new file is created.)

    """
    wb = lwb(xlsx)

    # Get sheet for global attributes
    ws = wb[GATT_SHEET_NAME]

    # Get indices of row for the gattr
    rows = get_row(ws, attname)

    if len(rows) == 0:
        logger.warning("There is no {0} attribute!".format(
                                        attname))
        return None

    # Update Cells
    for row in rows:
        # Remove gatt name
        ws.cell(row=row, column=1).value = None
        # Remove entry number
        ws.cell(row=row, column=2).value = None
        # Remove cdf dtype
        ws.cell(row=row, column=3).value = None
        # Remove entry value
        ws.cell(row=row, column=4).value = None

    return wb


def main():
    """Main program."""
    logger.info("This is the cdf.converter.tools.xlsx.gattrs module.")

# _________________ Main ____________________________
if (__name__ == "__main__"):
    main()
