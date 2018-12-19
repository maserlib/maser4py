#! /usr/bin/env python
# -*- coding: utf-8 -*-

"""Python module cdf.converter.tools.xlsx.gen."""

# ________________ IMPORT _________________________
# (Include here the modules to import, e.g. import sys)
import os.path as osp
import logging
from glob import glob

from openpyxl import load_workbook as lwb

__all__ = ["add_row", "get_row",
            "check_args", "get_item_list"]

# ________________ HEADER _________________________

# Mandatory
__version__ = ""
__author__ = "Xavier Bonnin"
__date__ = "2016-09-30"

# Optional
__license__ = ""
__credit__ = [""]
__maintainer__ = ""
__email__ = ""
__project__ = "MASER"
__institute__ = "LESIA, Observatoire de Paris, CNRS"
__changes__ = ""


# ________________ Global Variables _____________
# (define here the global variables)
logger = logging.getLogger(__name__)


# ________________ Class Definition __________
# (If required, define here classes)
class XlsxException(Exception):
    pass


# ________________ Global Functions __________
# (If required, define here gobal functions)

def check_args(sheet_name):
    """Decorator for arg checking and output file saving."""
    def decorated(func):
        def wrapper(*args, **kwargs):

            xlsx = args[0]

            if not osp.isfile(xlsx):
                msg = "Input file not found [{0}]!".format(xlsx)
                logger.error(msg)
                raise XlsxException(msg)
            else:
                wb = lwb(xlsx)

            shnames = wb.sheetnames
            if sheet_name not in shnames:
                msg = "Input file does not contain {0} sheet!".format(
                                            sheet_name)
                logger.error(msg)
                raise XlsxException(msg)

            if "overwrite" not in kwargs:
                kwargs["overwrite"] = False
            overwrite = kwargs["overwrite"]

            if "output" not in kwargs or kwargs["output"] is None:
                kwargs["output"] = xlsx
            output = kwargs["output"]

            wb = func(*args, **kwargs)
            if wb is None:
                return False

            if overwrite is False:
                nfile = len(glob(output + "*"))
                if nfile > 0:
                    output = output + "." + str(nfile)
            wb.save(output)

            return True
        return wrapper
    return decorated


def add_row(file, sheet, row,
            values=None,
            output=None,
            overwrite=True):
    """Insert a new row in the input Excel file.

    Positional arguments:
        file - Excel format file where the new row
               must be inserted
        sheet - Name of the sheet where the new row
                must be inserted
        row - rank of the row to insert (starting at 1)

    Optional keywords:
        values - List containing the Field values to be
                 inserted into the new
                 row (empty per default)
        output - Path of the output file where
                 the modified spreadsheet must
                 be saved. By default output is the
                 input file.
        overwrite - Overwrite existing output file

    Outputs:
        wb - Return the modifier workbook
    """
    if not osp.isfile(file):
        logger.error("Input Excel file not found [{0}]!".format(file))
        raise XlsxException

    if row < 1:
        logger.error("Row index must be greater or equal to 1!")
        raise XlsxException

    wb = lwb(file)
    sh = wb[sheet]

    max_row = sh.max_row
    max_col = sh.max_column

    if values is None:
        values = [None] * max_col
    else:
        max_col = len(values)

    if row == max_row:
        for col_num in range(max_col):
            sh.cell(row=max_row + 1,
                    column=col_num + 1).value = values[col_num]
    else:
        # Save the old row in a first buffer and insert the new one
        # Save also the row just below in a second buffer
        row_buff1 = [0] * max_col
        row_buff2 = [0] * max_col
        for col_num in range(0, max_col):
            row_buff1[col_num] = sh.cell(row=row, column=col_num + 1).value
            row_buff2[col_num] = sh.cell(row=row + 1, column=col_num + 1).value
            sh.cell(row=row, column=col_num + 1).value = values[col_num]

        # Shift rows by 1 in the bottom, starting from the
        # row to insert + 1, then copy row in buff1
        for row_num in range(row + 1, max_row + 2):
            for col_num in range(0, max_col):
                sh.cell(row=row_num, column=col_num + 1).value = \
                    row_buff1[col_num]
                row_buff1[col_num] = row_buff2[col_num]
                row_buff2[col_num] = sh.cell(row=row_num + 1,
                                            column=col_num + 1).value

    if output is not None:
        if osp.isfile(output) and not overwrite:
            logger.warning("Output file already exists [{0}]!".format(output))
        else:
            wb.save(output)
    return wb


def get_row(worksheet, name, column="A"):
    """get_row method.

    Check if an item exists or not in
    the given column of the given worksheet.
    If the item exists, then the method
    returns a list of the indice(s) of
    the item row(s).
    It returns an empty list otherwise.
    """
    # Check if the item alreay exists
    max_row = worksheet.max_row + 1
    rows = [i
        for i in range(1, max_row)
        if worksheet["{0}{1}".format(column, i)].value == name]

    return rows


def get_item_list(worksheet, column="A"):
    """Get the unique list of items for a given column."""
    items = list()
    max_row = worksheet.max_row + 1
    for i in range(1, max_row):
        val = worksheet["{0}{1}".format(column, i)].value
        if val not in items:
            items.append(worksheet["{0}{1}".format(column, i)].value)

    # Return item names (wihtout first header row)
    return items[1:]


def main():
    """Main program."""
    logger.info("This is the cdf.converter.tools.xlsx.gen module.")

# _________________ Main ____________________________
if (__name__ == "__main__"):
    main()
