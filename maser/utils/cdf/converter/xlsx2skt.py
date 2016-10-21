#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""xlsx2skt module.

Module to convert an Excel (.xlsx) file
into a CDF skeleton table (.skt).
"""

# ________________ HEADER _________________________

# ________________ IMPORT _________________________
# (Include here the modules to import, e.g. import sys)
import sys
import os
from datetime import datetime
import logging
import argparse

from openpyxl import load_workbook
from collections import OrderedDict

from ...toolbox import setup_logging, uniq, quote, truncate_str, insert_char
from ...._version import __version__

# ________________ HEADER _________________________

# ________________ Global Variables _____________
# (define here the global variables)
logger = logging.getLogger(__name__)

CURRENT_DATETIME = datetime.now()
ROW_LENGTH_MAX = 79
DEF_INDENT = " " * 16

# Sheets and columns to be found in the Excel file
SHEET_NAMES = {
    "header": [
        "CDF NAME",
        "DATA ENCODING",
        "MAJORITY",
        "FORMAT"
    ],
    "GLOBALattributes": [
        "Attribute Name",
        "Entry Number",
        "Data Type",
        "Value"
    ],
    "zVariables": [
        "Variable Name",
        "Data Type",
        "Number Elements",
        "Dims",
        "Sizes",
        "Record Variance",
        "Dimension Variances"
    ],
    "VARIABLEattributes": [
        "Variable Name",
        "Attribute Name",
        "Data Type",
        "Value"
    ],
    "Options": [
        "CDF_COMPRESSION",
        "CDF_CHECKSUM",
        "VAR_COMPRESSION",
        "VAR_SPARESERECORDS",
        "VAR_PADVALUE"
    ],
    "NRV": [
        "Variable Name",
        "Index",
        "Value"
    ]
}

# Available options
CDF_OPTION_NAMES = ["CDF_COMPRESSION", "CDF_CHECKSUM"]
VAR_OPTION_NAMES = ["VAR_COMPRESSION",
                    "VAR_SPARSERECORDS",
                    "VAR_PADVALUE"]

# CDF skeleton headers format
HEADER_BOARD = "! Variables     G.Attributes     " + \
    "V.Attributes     Records     Dims     Sizes\n"
HEADER_BOARD += "! ---------     ------------  " + \
    "   ------------     -------     ----     -----"
HEADER_SPACE = DEF_INDENT

GLOBAL_BOARD = "! Attribute    Entry        Data\n"
GLOBAL_BOARD += "! Name         Number       Type   Value\n"
GLOBAL_BOARD += "! ---------    ------       ----   -----"

VARIABLE_BOARD = "! Variable    Data     Number " + \
    "                     Record       Dimension\n"
VARIABLE_BOARD += "! Name        Type     Elements" + \
    "   Dims    Sizes    Variance     Variances\n"
VARIABLE_BOARD += "! --------    ----     -------- " + \
    "  ----    -----    --------     ---------"

VATTRS_BOARD = "  ! Attribute    Data\n"
VATTRS_BOARD += "  ! Name         Type   Value\n"
VATTRS_BOARD += "  ! --------     ----   -----"


# ________________ Class Definition __________
# (If required, define here classes)
class Xlsx2skt:
    """Class to convert a formatted Excel file into a CDF skeleton table."""

    def __init__(self, xlsx_file,
                 skt_file=None,
                 output_dir=None,
                 overwrite=False,
                 ignore_none=False,
                 auto_pad=False,
                 verbose=True,
                 debug=False,
                 quiet=False):
        """__init__ method."""
        self.xlsx_file = xlsx_file
        self.overwrite = overwrite
        self.ignore = ignore_none
        self.auto_pad = auto_pad

        self.cdf_items = {}

        if skt_file is None:
            skt_file = os.path.splitext(self.xlsx_file)[0] + ".skt"

        if output_dir is None:
            output_dir = os.path.basename(xlsx_file)
        else:
            skt_file = os.path.join(output_dir, os.path.basename(skt_file))

        self.skt_file = skt_file

    # Setup the logging
        setup_logging(
            filename=None, quiet=quiet,
            verbose=verbose,
            debug=debug)

    def _parse_xlsx(self):
        """Parse the Excel 2007 format file."""
        xlsx = self.xlsx_file

        if not os.path.isfile(xlsx):
            logger.error("Cannot find Excel file called %s!", xlsx)
            sys.exit(-1)

        if os.path.splitext(xlsx)[1] != ".xlsx":
            logger.error("Invalid input Excel format!")
            sys.exit(-1)

        logger.info("Parsing %s file...", xlsx)
        wkbk = load_workbook(xlsx, read_only=True)
        sheet_names = wkbk.get_sheet_names()

        if ("rVariables" in sheet_names) or ("variables" in sheet_names):
            logger.warning("rVariable type is not supported!")

        sheets = dict()
        for shtn in SHEET_NAMES:
            logger.info("Loading %s sheet...", shtn)
            if shtn not in sheet_names:
                logger.error("Missing %s sheet in the input Excel file!", shtn)
                sys.exit(-1)
            else:
                wksht = wkbk[shtn]

                ncol = wksht.max_column
                nrow = wksht.max_row
                if ncol is None or nrow is None:
                    wksht.calculate_dimension(force=True)
                    ncol = wksht.max_column
                    nrow = wksht.max_row

                ncol += 1
                nrow += 1

                sheet_data = OrderedDict()
                for i, row in enumerate(wksht.rows):
                    cells_i = []
                    for cell in row:
                        if i == 0:
                            sheet_data[cell.value] = []
                        else:
                            cells_i.append(cell.value)
                    if i != 0:
                        for k, key in enumerate(sheet_data):
                            sheet_data[key].append(cells_i[k])

                sheets[shtn] = sheet_data

        self.cdf_items["GLOBALattributes"] = \
            uniq(sheets["GLOBALattributes"]["Attribute Name"],
                 not_none=True)
        self.cdf_items["VARIABLEattributes"] = \
            uniq(sheets["VARIABLEattributes"]["Attribute Name"],
                 not_none=True)
        self.cdf_items["zVariables"] = \
            uniq(sheets["zVariables"]["Variable Name"],
                 not_none=True)
        logger.info("%i GLOBAL attributes returned",
                  len(self.cdf_items["GLOBALattributes"]))
        logger.info("%i Variable attributes returned",
                  len(self.cdf_items["VARIABLEattributes"]))
        logger.info("%i zVariables returned",
                  len(self.cdf_items["zVariables"]))

        return sheets

    def _build_skt(self, xlsx_sheets):
        """Build the CDF skeleton table content using the Excel data."""
        logger.info("Building CDF skeleton table body... ")

        skt_name = os.path.splitext(os.path.basename(self.skt_file))[0]
        xlsx_name = os.path.basename(self.xlsx_file)

        file_header = "!Skeleton table for the \"" + skt_name + "\" CDF.\n"
        file_header += "!Generated: " + \
            CURRENT_DATETIME.strftime("%Y-%m-%d %H:%M:%S") + "\n"
        file_header += "!Skeleton table created by xlsx2skt.py V" + \
            __version__ + "\n"
        file_header += "!Skeleton table created from " + xlsx_name + "\n"

        skt_header = self._build_header(xlsx_sheets["header"],
                                       xlsx_sheets["Options"])
        skt_global = self._build_global(xlsx_sheets["GLOBALattributes"])
        skt_vattrs = self._build_vattributes()
        skt_zvars = self._build_zvariables(xlsx_sheets["zVariables"],
                                          xlsx_sheets["VARIABLEattributes"],
                                          xlsx_sheets["Options"],
                                          xlsx_sheets["NRV"],
                                          ignore_none=self.ignore,
                                          auto_pad=self.auto_pad)

        skt_body = "\n".join([file_header, "", skt_header, "",
                              skt_global, "", skt_vattrs, "",
                              skt_zvars, "", "#end"])

        return skt_body

    def _write_skt(self, skt_body):
        """Write the CDF skeleton table file."""
        skt = self.skt_file

        if os.path.splitext(skt)[1] != ".skt":
            logger.info(
                ".skt extension will be automatically appended to %s",
                skt)
            skt = skt + ".skt"

        if not (self.overwrite) and (os.path.isfile(skt)):
            logger.warning("%s already exits!", skt)
            return skt

        logger.info("Writing %s...", skt)

        with open(skt, "w") as filew:
            filew.write(skt_body)

        if os.path.isfile(skt):
            logger.info(skt + " has been saved correctly")
            return skt
        else:
            logger.error(skt + " has not been saved correctly!")
            return None

    @classmethod
    def convert(cls, *args, **kwargs):
        """Run the complete xlsx to skt conversion process."""
        xlsx2skt = cls(*args, **kwargs)

        xlsx_sheets = xlsx2skt._parse_xlsx()
        skt_body = xlsx2skt._build_skt(xlsx_sheets)
        skt_path = xlsx2skt._write_skt(skt_body)

        if skt_path:
            return True
        else:
            return False

    def _build_header(self, header_sheet, options_sheet):
        """Build the CDF skeleton table header part."""
        logger.info("Building skeleton table header...")

        header_body = ["#header", ""]

        for col in SHEET_NAMES["header"]:
            if col in header_sheet:
                val = header_sheet[col]
                header_body.append(HEADER_SPACE + col + ": " + val[0])
            else:
                logger.error(col +
                             " column is missing in the \"header\" sheet!")
                sys.exit(-1)

        header_body.append("")
        header_body.append(HEADER_BOARD)

        nglobal = len(self.cdf_items["GLOBALattributes"])
        nvattr = len(self.cdf_items["VARIABLEattributes"])
        nvar = len(self.cdf_items["zVariables"])

        header_board_info = ["    0/" + str(nvar),
                             str(nglobal), str(nvattr), "0/z", "0"]
        header_board_info = "            ".join(header_board_info)
        header_body.append(header_board_info)

        header_opt_info = "\n"
        for opt in CDF_OPTION_NAMES:
            if opt in options_sheet:
                header_opt_info += "!" + opt + ": " + \
                    options_sheet[opt][0] + "\n"
        header_body.append(header_opt_info)

        header_body = "\n".join(header_body)

        logger.debug(header_body)

        return header_body

    def _build_global(self, global_sheet):
        """Build the CDF skeleton table GLOBALattributes part."""
        logger.info("Building skeleton table global attributes section..")

        global_body = ["#GLOBALattributes", ""]

        global_body.append(GLOBAL_BOARD)

        if global_sheet["Attribute Name"][0] is None:
            logger.error("First Global attribute name must not be null!")
            sys.exit(-1)
        else:
            last_valid_attr = global_sheet["Attribute Name"][0]

        nindent = DEF_INDENT
        new_entry = ""
        for i, attr in enumerate(global_sheet["Attribute Name"]):

            if attr is None:
                logging.warning("Attribute #%i is null!", i)
                continue

            if last_valid_attr != attr:
                new_entry += " .\n"
            global_body.append(new_entry)

            enum_i = global_sheet["Entry Number"][i]
            if enum_i is None:
                logger.error("Attribute \"%s\" Entry Number is empty!", attr)
                sys.exit(-1)
            else:
                enum_i = str(enum_i)

            dtype_i = global_sheet["Data Type"][i]
            if dtype_i is None:
                logger.error("Attribute \"%s\" Data Type is empty!", attr)
                sys.exit(-1)
            else:
                dtype_i = str(dtype_i)

            if ("CDF_CHAR" in dtype_i or
                    "CDF_UCHAR" in dtype_i):
                ischar = True
            else:
                ischar = False

            value_i = str(global_sheet["Value"][i])
            value_i = quote(value_i, unquote=True)

            if (value_i is None or
               value_i.lower() == "none" or
               value_i == ""):
                logger.warning("Attribute \"%s\" value is empty!", attr)
                value_i = " "

            if int(enum_i) == 1:
                new_entry = "  " + quote(attr) + "        "
            else:
                new_entry = " " * (nindent + 8)

            new_entry += enum_i + ":  " + dtype_i + "    { "

            # Remove any break line
            value_i = value_i.replace("\n", " ")

            # If attribute is a string, then cut it if it
            # has a too long length
            if ischar:
                value_i = truncate_str(value_i,
                                    int(ROW_LENGTH_MAX / 3),
                                    gap=(" " * (len(new_entry) + 12)),
                                    min_length=6)

                new_entry += quote(value_i) + " }"
            else:
                new_entry += value_i + " }"

            if len(new_entry) > ROW_LENGTH_MAX and int(enum_i) == 1:
                nindent = len(attr)
                new_entry = insert_char(new_entry, "\n" + " " * nindent,
                                        nindent + 4)
            elif int(enum_i) == 1:
                nindent = len(attr) + 14

            last_valid_attr = attr

        new_entry += " .\n"
        global_body.append(new_entry)
        global_body = "\n".join(global_body)

        logger.debug(global_body)

        return global_body

    def _build_vattributes(self):
        """Build the list of variable attributes."""
        logger.info("Building skeleton table variable attributes section...")

        vattrs_body = ["#VARIABLEattributes", ""]

        for vattr in self.cdf_items["VARIABLEattributes"]:
            vattrs_body.append("  " + quote(vattr))

        vattrs_body = "\n".join(vattrs_body)

        logger.debug(vattrs_body)

        return vattrs_body

    def _build_zvariables(self, zvars_sheet, vattrs_sheet,
                         options_sheet, nrv_sheet,
                         ignore_none=False,
                         auto_pad=True):
        """Build the CDF skeleton table.

        VARIABLEattributes and zVariables parts.
        """
        logger.info("Building skeleton table zvariable section...")

        zvar_body = ["#variables", ""]
        zvar_body.extend(["!No rVariables.", ""])
        zvar_body.extend(["#zVariables", ""])

        for i, zvar in enumerate(zvars_sheet["Variable Name"]):

            if zvar is None:
                if ignore_none:
                    logger.warning("Current zVariable is NoneType, skipping!")
                    continue
                else:
                    logger.error("ERROR: Current zVariable is NoneType!")
                    raise TypeError

            zvar_body.append(VARIABLE_BOARD)

            # New zVariable entry
            zvar_body.append("")

            dtype_i = str(zvars_sheet["Data Type"][i])
            nelem_i = str(zvars_sheet["Number Elements"][i])
            dims_i = str(zvars_sheet["Dims"][i])
            sizes_i = str(zvars_sheet["Sizes"][i])
            recvar_i = str(zvars_sheet["Record Variance"][i])
            dimvars_i = str(zvars_sheet["Dimension Variances"][i])

            logger.debug("  " + quote(zvar) + "    " + dtype_i +
                      "     " + nelem_i + "     " + dims_i + "     " +
                      sizes_i + "     " + recvar_i + "     " + dimvars_i)

            if dtype_i == "None":
                logger.error("Wrong Data Type for %s!", zvar)
                sys.exit(-1)
            if nelem_i == "None":
                logger.error("Wrong Number Elements for %s!", zvar)
                sys.exit(-1)
            if dims_i == "None":
                logger.error("Wrong Dims for %s!", zvar)
                sys.exit(-1)

            if sizes_i == "None":
                sizes_i = ""
            if dimvars_i == "None":
                dimvars_i = ""

            zvar_entry = "  " + quote(zvar) + "    " + dtype_i + \
                "     " + nelem_i + "     " + dims_i + "     " + \
                sizes_i + "     " + recvar_i + "     " + dimvars_i

            if len(zvar_entry) > ROW_LENGTH_MAX:
                zvar_entry = insert_char(zvar_entry, "\n    ", len(zvar) + 4)

            zvar_body.append(zvar_entry)

            var_opt_info = "\n"
            for opt in VAR_OPTION_NAMES:
                if opt in options_sheet:
                    if opt == "VAR_PADVALUE" and auto_pad:
                        options_sheet[opt][0] = assign_pad(dtype_i)
                    var_opt_info += "! " + opt + ": " + \
                        options_sheet[opt][0] + "\n"
            zvar_body.append(var_opt_info)

            # Add variable attributes info
            zvar_body.append(VATTRS_BOARD)

            vattr_entry = ""
            for j, vattr_var in enumerate(vattrs_sheet["Variable Name"]):

                if vattr_var == zvar:
                    vattr_entry += "\n"
                    vattr_name = quote(str(vattrs_sheet["Attribute Name"][j]))
                    vattr_dtype = str(vattrs_sheet["Data Type"][j])
                    if vattr_dtype == "None":
                        logger.error("Wrong Data Type for the " +
                                    "attribute %s " +
                                    "of the variable %s !", vattr_name, zvar)
                        raise TypeError

                    if vattrs_sheet["Value"][j] is None:
                        vattr_value = ""
                    else:
                        vattr_value = str(vattrs_sheet["Value"][j])

                    vattr_entry_j = "   " + vattr_name + "    " \
                        + vattr_dtype + "     { "

                    if vattr_dtype == "CDF_CHAR":
                        if len(vattr_value) == 0:
                            vattr_value = " "

                        vattr_value = vattr_value.replace("\n", " ")
                        gap = " " * (2 * len(vattr_entry_j) - 1)
                        vattr_value = truncate_str(vattr_value,
                                                   int(ROW_LENGTH_MAX / 3),
                                                   gap=gap,
                                                   min_length=6)
                        vattr_value = quote(vattr_value)

                    vattr_entry += vattr_entry_j + vattr_value + " }"

            vattr_entry += " .\n"
            zvar_body.append(vattr_entry)

            # Add NRV
            nrv_body = ""
            for k, nrv_k in enumerate(nrv_sheet["Variable Name"]):
                if zvar == nrv_k:
                    idx_k = str(nrv_sheet["Index"][k])
                    val_k = str(nrv_sheet["Value"][k])
                    if (idx_k == "") or (idx_k == "None"):
                        logger.error("Wrong NRV index for %s!", nrv_k)
                        raise TypeError
                    if (dtype_i == "CDF_CHAR" or
                    dtype_i == "CDF_UCHAR"):
                        val_k = "{ " + quote(val_k) + " }"
                    nrv_body += "    [ " + idx_k + " ] = " + \
                        val_k + "\n"

            if len(nrv_body) == 0:
                nrv_body = "  ! RV values were not requested.\n"
            else:
                nrv_body = "  ! NRV values follow...\n\n" + nrv_body

            zvar_body.append(nrv_body)

        zvar_body = "\n".join(zvar_body)
        return zvar_body


# ________________ Global Functions __________
def assign_pad(data_type):
    """VAR_PADVALUE auto assign.

    Automatically assigns VAR_PADVALUE
    depending to the input data_type
    """
    dtype = data_type.upper()

    if "EPOCH" in dtype:
        return "01-Jan-0000 00:00:00.000"
    elif "TT2000" in dtype:
        return "0000-01-01T00:00:00.000000000"
    elif ("INT" in dtype) or ("BYTE" in dtype):
        return "0"
    elif ("FLOAT" in dtype) or ("REAL" in dtype) or ("DOUBLE" in dtype):
        return "0.0"
    elif "CHAR" in dtype:
        return "\" \""
    else:
        return "None"


def main():
    """xlsx2skt main program."""
    parser = argparse.ArgumentParser(
        description='CDF converter main modulea Excel 2007 ' +
        'format file into a CDF skeleton table',
        add_help=True)
    parser.add_argument('excel', nargs=1,
                        help='Excel 2007 format file (.xlsx)')
    parser.add_argument('-s', '--skeleton', nargs='?',
                        default=None,
                        help='Output CDF skeleton table (.skt)')
    parser.add_argument('-o', '--output_dir', nargs='?',
                        default=None,
                        help='Path of the output directory')
    parser.add_argument('-O', '--Overwrite', action='store_true',
                        help='Overwrite existing output files')
    parser.add_argument('-V', '--Verbose', action='store_true',
                        help='Verbose mode')
    parser.add_argument('-D', '--Debug', action='store_true',
                        help='Debug mode')
    parser.add_argument('-Q', '--Quiet', action='store_true',
                        help='Quiet mode')
    parser.add_argument('-I', '--Ignore_none', action='store_true',
                        help='Ignore NoneType zVariables')
    parser.add_argument('-A', '--Auto_pad', action='store_true',
                        help='Value of !VAR_PADVALUE ' +
                        'is automatically assigned')
    args = parser.parse_args()

    Xlsx2skt.convert(args.excel[0],
                  skt_file=args.skeleton,
                  output_dir=args.output_dir,
                  overwrite=args.Overwrite,
                  ignore_none=args.Ignore_none,
                  auto_pad=args.Auto_pad,
                  verbose=args.Verbose,
                  debug=args.Debug,
                  quiet=args.Quiet)

# _________________ Main ____________________________
if __name__ == "__main__":
    main()